from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from flask import request, render_template, redirect, session, Flask
import datetime
import matplotlib.pyplot as plt
import numpy as np
from pytz import timezone



def append_data(type, cost, description):
    wb = load_workbook('finances.xlsx')

    ws = wb.active
    date=datetime.datetime.now()
    
    my_timezone=timezone('US/Pacific')

    date = my_timezone.localize(date)
    date = date.astimezone(my_timezone)
    ws.append([str(type), str(cost), description, date.strftime("%B %d, %Y")])
    wb.save('finances.xlsx')
    



def show_data():
    wb = load_workbook('finances.xlsx')
    ws = wb.active
   
    food = 0
    videogames = 0
    computer = 0
    other = 0
    savings = 0
    data=[]
    count=1
    for row in ws.values:


        row_data={}
        row_data["type"] = row[0]
        row_data["cost"] = "{0:.2f}".format(float(row[1]))
        row_data['description'] = row[2]
        row_data["date"] = row[3]
        row_data["row"] = count
        if row[0] == 'Food':
            food += float(row[1])
        if row[0] == 'Videogames':
            videogames += float(row[1])
        if row[0] == 'Computer':
            computer += float(row[1])
        if row[0] == 'Other':
            other += float(row[1])
        if row[0] == 'Saving':
            savings += float(row[1])
        

        data.append(row_data)
        count+=1

    
    saved = savings
    spent = food + videogames + computer + other
    total = spent + savings
    costs=[]
    mylabels=[]
    colours=[]
    if food > 0:
        costs.append(food)
        mylabels.append("Food")
        colours.append('orchid')
    
    if videogames > 0:
        costs.append(videogames)
        mylabels.append("Videogames")
        colours.append("coral")
    
    if computer > 0:
        costs.append(computer)
        mylabels.append("Computer")
        colours.append("gold")

    if other > 0:
        costs.append(other)
        mylabels.append("Other")
        colours.append('limegreen')

    if savings > 0:
        costs.append(savings)
        mylabels.append("Savings")
        colours.append('deepskyblue')

    plt.switch_backend('agg')
    plt.pie(costs, labels=mylabels, autopct=lambda p: '${:.2f}'.format(p * sum(costs) / 100), colors=colours)
    plt.legend(title = 'Costs')
    plt.title("Finance Pie Graph")
    plt.savefig('static/img/piechart.png')

    fig, ax = plt.subplots()
    ax.bar(mylabels, costs, color=colours)
    ax.set_ylabel("Total Cost ($)")
    ax.set_title("Finance Bar Graph")
 
    plt.savefig('static/img/barchart.png')
  
    total = '{0:.2f}'.format(total)
    saved = '{0:.2f}'.format(saved)
    spent = '{0:.2f}'.format(spent)
  
    return data, total, saved, spent

def delete_row(row):
    wb = load_workbook('finances.xlsx')
    ws = wb.active
    ws.delete_rows(int(row))
    wb.save('finances.xlsx')


def reset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Finances"
    wb.save('finances.xlsx')